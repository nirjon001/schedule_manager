import openpyxl
import re
from collections import OrderedDict

DAY_ORDER = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']

TIME_SLOTS = [
    '8:00AM', '8:30AM', '9:00AM', '9:30AM', '10:00AM', '10:10AM', '10:30AM',
    '11:00AM', '11:30AM', '11:50AM', '12:00PM', '12:10PM', '12:30PM',
    '1:00PM', '1:20PM', '1:30PM', '2:00PM', '2:30PM', '3:00PM', '3:30PM',
    '4:00PM', '4:30PM', '4:50PM', '5:00PM', '5:30PM', '6:00PM', '6:20PM', '6:30PM', '6:50PM',
    '7:00PM', '7:30PM', '8:00PM',
]


def parse_time_day(time_day_str):
    if not time_day_str:
        return []
    time_day_str = str(time_day_str).strip()
    match = re.match(r'([ASMTWRF]+)\s+(\d{1,2}:\d{2}(?:AM|PM))-(\d{1,2}:\d{2}(?:AM|PM))', time_day_str)
    if not match:
        return []
    days_str, start_time, end_time = match.groups()

    mapping = {}
    for ch in days_str:
        if ch == 'A':
            mapping['A'] = 'Saturday'
        elif ch == 'S':
            mapping['S'] = 'Sunday'
        elif ch == 'M':
            mapping['M'] = 'Monday'
        elif ch == 'T':
            mapping['T'] = 'Tuesday'
        elif ch == 'W':
            mapping['W'] = 'Wednesday'
        elif ch == 'R':
            mapping['R'] = 'Thursday'
        elif ch == 'F':
            mapping['F'] = 'Friday'

    result = []
    for day_name in mapping.values():
        result.append({'day': day_name, 'start': start_time, 'end': end_time})
    return result


def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    student_info = {'name': '', 'id': '', 'semester': ''}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            val = cell.value
            if not val:
                continue
            val_str = str(val).strip()
            if val_str == 'Name:':
                for c in row:
                    cv = str(c.value or '').strip()
                    if cv != 'Name:' and cv and len(cv) > 3:
                        student_info['name'] = cv
                        break
            elif val_str == 'ID#':
                id_cell = ws.cell(row=cell.row, column=cell.column + 2)
                if id_cell.value:
                    student_info['id'] = str(id_cell.value).strip()

    semester_str = ''
    for row in ws.iter_rows(min_row=14, max_row=15, values_only=False):
        for cell in row:
            val = str(cell.value or '').strip()
            if any(s in val for s in ['Spring', 'Summer', 'Fall']):
                semester_str = val
                break

    student_info['semester'] = semester_str

    courses = []
    in_course_section = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        row_vals = {c.column: c.value for c in row}

        if any(str(row_vals.get(c, '') or '').strip() == 'Course(s)' for c in [3, 4, 5, 6, 7]):
            in_course_section = True
            continue

        if not in_course_section:
            continue

        course_cell = row_vals.get(3)
        if not course_cell:
            course_cell = row_vals.get(4) or row_vals.get(5)
        if not course_cell:
            continue

        course_str = str(course_cell).strip()
        if not re.match(r'^[A-Z]{3}\d{3}', course_str) and not re.match(r'^[A-Z]{3}\d{3}\s+Lab', course_str):
            continue

        time_day_val = str(row_vals.get(53, '') or '').strip()
        room_val = str(row_vals.get(65, '') or '').strip()
        sec_val = str(row_vals.get(17, '') or '').strip()

        if not time_day_val or not re.match(r'[ASMTWRF]+\s+\d{1,2}:\d{2}(?:AM|PM)-\d{1,2}:\d{2}(?:AM|PM)', time_day_val):
            continue

        schedules = parse_time_day(time_day_val)
        for sched in schedules:
            courses.append({
                'course': course_str,
                'sec': sec_val,
                'day': sched['day'],
                'start': sched['start'],
                'end': sched['end'],
                'room': room_val,
            })

    for c in courses:
        code = c['course']
        if code.endswith(' Lab') and not c.get('sec'):
            parent = code[:-4]
            for pc in courses:
                if pc['course'] == parent and pc.get('sec'):
                    c['sec'] = pc['sec']
                    break

    return student_info, courses


def organize_by_day(courses):
    schedule = OrderedDict()
    for day in DAY_ORDER:
        schedule[day] = []
    for c in courses:
        if c['day'] in schedule:
            schedule[c['day']].append(c)
    for day in schedule:
        schedule[day].sort(key=lambda x: x['start'])
    schedule = OrderedDict((k, v) for k, v in schedule.items() if v)
    return schedule


def sort_time_key(course):
    return TIME_SLOTS.index(course['start']) if course['start'] in TIME_SLOTS else 999


def build_grid(schedule):
    active_days = list(schedule.keys())
    time_ranges = OrderedDict()
    for day, courses in schedule.items():
        for c in courses:
            key = (c['start'], c['end'])
            if key not in time_ranges:
                time_ranges[key] = []
    time_ranges = OrderedDict(sorted(time_ranges.items(), key=lambda x: sort_time_key({'start': x[0][0]})))

    grid_rows = []
    for (start, end) in time_ranges:
        row_data = {'start': start, 'end': end, 'courses': {}}
        for day in active_days:
            row_data['courses'][day] = None
        for day, courses in schedule.items():
            for c in courses:
                if c['start'] == start and c['end'] == end:
                    row_data['courses'][day] = c
        grid_rows.append(row_data)

    return active_days, grid_rows
